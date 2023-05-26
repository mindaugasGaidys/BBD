import pandas as pd
from sklearn.model_selection import train_test_split
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import re
import contractions
import openpyxl

data = pd.read_excel("Data/Datasets/NB.xlsx")

stop_words = set(stopwords.words('english'))
lemmatizer = WordNetLemmatizer()

def nb_data_preprocess(text):
    text = str(text)
    text = contractions.fix(text)
    text = re.sub(r'\W', ' ', text)
    text = text.lower()
    text = re.sub(r'\s+[a-zA-Z]\s+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    words = text.split()
    words = [lemmatizer.lemmatize(word) for word in words if word not in stop_words]
    return ' '.join(words)

data['text_cleaned'] = data['text'].apply(nb_data_preprocess)
data = data.drop(columns=['text'])
train_data, test_data = train_test_split(data, test_size=0.2, random_state=42)

excel_file = "Data/Datasets/NB.xlsx"
wb = openpyxl.load_workbook(excel_file)
if 'NB_train' in wb.sheetnames:
    wb.remove(wb['NB_train'])
if 'NB_test' in wb.sheetnames:
    wb.remove(wb['NB_test'])
wb.save(excel_file)

with pd.ExcelWriter("Data/Datasets/NB.xlsx", engine='openpyxl', mode='a') as writer:
    train_data.to_excel(writer, sheet_name='NB_train', index=False)
    test_data.to_excel(writer, sheet_name='NB_test', index=False)