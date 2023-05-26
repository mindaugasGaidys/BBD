from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options

chrome_options = Options()
chrome_options.add_argument("--disable-images")
chrome_options.add_argument("--blink-settings=imagesEnabled=false")

driver = webdriver.Chrome(options=chrome_options)
driver.get("https://news.sky.com/story/ukraine-russia-war-latest-putin-assassination-attempt-as-kremlin-residence-attacked-by-drones-moscow-claims-12541713")
time.sleep(5)
while True:
    section_element = driver.find_element(By.CLASS_NAME, "ncpost-list-container")
    child_elements = section_element.find_elements(By.XPATH, "./*")
    last_datetime = child_elements[-1].find_element(By.XPATH, ".//div[@class='ncpost-header']//time").get_attribute(
        "datetime")[:10]
    if last_datetime == "2023-01-01":
        break
    try:
        load_more_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Load More')]")
        load_more_button.click()
        time.sleep(2)
        print("Clicked")
    except NoSuchElementException:
        print("Cant find button")
        break

workbook = openpyxl.load_workbook("Data\\Datasets\\NewsArticles.xlsx")
print("workbook loaded")

worksheet = workbook["Sheet1"]

section_element = driver.find_element(By.CLASS_NAME, "ncpost-list-container")
print("content part found")
child_elements = section_element.find_elements(By.XPATH, "./*")
print("articles found")
last_row = worksheet.max_row
count=0
for row, child_element in enumerate(child_elements, start=last_row+1):
    datetime_element = child_element.find_element(By.XPATH, ".//div[@class='ncpost-header']//time")
    datetime = datetime_element.get_attribute("datetime")[:10]
    
    container_element = child_element.find_element(By.XPATH, ".//div[@class='ncpost-container']")
    try:
        content_element = container_element.find_element(By.XPATH, ".//div[@class='ncpost-content']")
        paragraph_elements = content_element.find_elements(By.XPATH, ".//p")
        paragraphs = "\n".join([paragraph.text for paragraph in paragraph_elements])
    except NoSuchElementException:
        continue
    
    worksheet.cell(row=row, column=1, value=datetime)
    worksheet.cell(row=row, column=2, value=paragraphs)
    count+=1
    print(count+ " Articles written to excel file")

workbook.save("Data\\Datasets\\NewsArticles.xlsx")
print("Articles saved")
driver.quit()
