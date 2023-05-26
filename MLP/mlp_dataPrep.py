import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
import contractions
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer
import re

data = pd.read_excel('Data/Datasets/MLP.xlsx', sheet_name='Sheet1', engine='openpyxl')

data['text_expanded'] = data['text'].apply(lambda x: contractions.fix(str(x)))
stop_words = set(stopwords.words('english'))

lemmatizer = WordNetLemmatizer()
def preprocess_text(text):
    text = text.lower()
    text = re.sub(r'[^\w\s]', '', text)
    tokenized_text = word_tokenize(text)
    tokenized_text = [lemmatizer.lemmatize(word) for word in tokenized_text if word not in stop_words]
    return ' '.join(tokenized_text)

data['text_cleaned'] = data['text_expanded'].apply(preprocess_text)
data['text_cleaned'].fillna('', inplace=True)
data = data.dropna(subset=['category'])

X_train, X_test, y_train, y_test = train_test_split(data['text_cleaned'], data['category'], test_size=0.2, random_state=42, stratify=data['category'])

book = load_workbook('Data/Datasets/MLP.xlsx')

for sheet_name in ['MLP_train', 'MLP_test']:
    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])

book.save('Data/Datasets/MLP.xlsx')

train_data = pd.DataFrame(X_train)
train_data.columns = ['text_cleaned']
train_data['category'] = y_train.values
with pd.ExcelWriter('Data/Datasets/MLP.xlsx', engine='openpyxl', mode='a') as writer:
    train_data.to_excel(writer, sheet_name='MLP_train', index=False)

test_data = pd.DataFrame(X_test)
test_data.columns = ['text_cleaned']
test_data['category'] = y_test.values
with pd.ExcelWriter('Data/Datasets/MLP.xlsx', engine='openpyxl', mode='a') as writer:
    test_data.to_excel(writer, sheet_name='MLP_test', index=False)
