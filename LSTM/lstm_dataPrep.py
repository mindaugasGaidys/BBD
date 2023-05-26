import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
import contractions
from nltk.corpus import stopwords


data = pd.read_excel('Data\Datasets\LSTM.xlsx', sheet_name='AllData', engine='openpyxl')


data['text_expanded'] = data['text'].apply(lambda x: contractions.fix(str(x)))
stop_words = set(stopwords.words('english'))
data['text_cleaned'] = data['text_expanded'].apply(lambda x: ' '.join([word.lower() for word in x.split() if word.lower() not in stop_words]))
data = data.dropna(subset=['text_cleaned', 'category'])

X_train, X_test, y_train, y_test = train_test_split(data['text_cleaned'], data['category'], test_size=0.2, random_state=42, stratify=data['category'])

book = load_workbook('Data\Datasets\LSTM.xlsx')

for sheet_name in ['LSTM_train', 'LSTM_test']:
    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])

book.save('Data\Datasets\LSTM.xlsx')

train_data = pd.DataFrame(X_train)
train_data.columns = ['text_cleaned']
train_data['category'] = y_train.values
with pd.ExcelWriter('Data\Datasets\LSTM.xlsx', engine='openpyxl', mode='a') as writer:
    train_data.to_excel(writer, sheet_name='LSTM_train', index=False)

test_data = pd.DataFrame(X_test)
test_data.columns = ['text_cleaned']
test_data['category'] = y_test.values
with pd.ExcelWriter('Data\Datasets\LSTM.xlsx', engine='openpyxl', mode='a') as writer:
    test_data.to_excel(writer, sheet_name='LSTM_test', index=False)